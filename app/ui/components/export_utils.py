"""
Export utilities for BitCraft Companion.

Provides functions to export inventory and other data to Excel format.
"""

import logging
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime


def export_inventory_to_excel(inventory_data, file_path):
    """
    Export inventory data to Excel format.
    
    Args:
        inventory_data: List of inventory items from ClaimInventoryTab.all_data
        file_path: Path where to save the Excel file
        
    Returns:
        bool: True if export succeeded, False otherwise
    """
    try:
        if not inventory_data:
            logging.warning("No inventory data provided for export")
            return False
            
        # Create workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Claim Inventory"
        
        # Define headers to match the real inventory data format
        headers = ["Name", "Tier", "Quantity", "Tag", "Containers"]
            
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            
        # Write data rows
        for row_idx, item in enumerate(inventory_data, 2):
            if isinstance(item, dict):
                # Extract values using the correct field names from inventory tab
                name = item.get("name", "")
                tier = item.get("tier", "")
                quantity = item.get("quantity", "")
                tag = item.get("tag", "")
                containers = item.get("containers", {})
                
                # Format containers dict into readable string
                if isinstance(containers, dict):
                    if len(containers) == 0:
                        containers_str = ""
                    elif len(containers) == 1:
                        # Single container: "Container Name: 123"
                        container_name, container_qty = next(iter(containers.items()))
                        containers_str = f"{container_name}: {container_qty}"
                    else:
                        # Multiple containers: "Container1: 50, Container2: 25"
                        container_strs = [f"{name}: {qty}" for name, qty in containers.items()]
                        containers_str = ", ".join(container_strs)
                else:
                    containers_str = str(containers) if containers else ""
                
                # Write the row data
                row_values = [name, tier, quantity, tag, containers_str]
                for col_idx, value in enumerate(row_values, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
            else:
                # Fallback for non-dict items
                logging.warning(f"Unexpected item format: {type(item)} - {item}")
                sheet.cell(row=row_idx, column=1, value=str(item))
        
        # Auto-adjust column widths
        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        # Save workbook
        workbook.save(file_path)
        logging.info(f"Successfully exported {len(inventory_data)} inventory items to: {file_path}")
        return True
        
    except Exception as e:
        logging.error(f"Error exporting inventory data to Excel: {e}")
        return False


def export_multiple_sheets_to_excel(sheets_data, file_path):
    """
    Export multiple sheets of data to Excel format.
    
    Args:
        sheets_data: Dictionary with sheet_name -> data mappings
        file_path: Path where to save the Excel file
        
    Returns:
        bool: True if export succeeded, False otherwise
    """
    try:
        if not sheets_data:
            logging.warning("No sheets data provided for export")
            return False
            
        # Create workbook
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        
        # Create sheets
        for sheet_name, sheet_data in sheets_data.items():
            if not sheet_data:
                continue
                
            # Create sheet
            sheet = workbook.create_sheet(title=sheet_name)
            
            # Determine headers
            if isinstance(sheet_data, list) and len(sheet_data) > 0:
                if isinstance(sheet_data[0], dict):
                    headers = list(sheet_data[0].keys())
                elif hasattr(sheet_data[0], '__iter__') and not isinstance(sheet_data[0], str):
                    headers = [f"Column_{i+1}" for i in range(len(sheet_data[0]))]
                else:
                    headers = ["Value"]
            else:
                headers = ["Value"]
                
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                
            # Write data
            for row_idx, item in enumerate(sheet_data, 2):
                if isinstance(item, dict):
                    for col_idx, header in enumerate(headers, 1):
                        value = item.get(header, "")
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                elif hasattr(item, '__iter__') and not isinstance(item, str):
                    for col_idx, value in enumerate(item, 1):
                        if col_idx <= len(headers):
                            sheet.cell(row=row_idx, column=col_idx, value=value)
                else:
                    sheet.cell(row=row_idx, column=1, value=item)
            
            # Auto-adjust column widths
            for column_cells in sheet.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)
        
        # Save workbook
        workbook.save(file_path)
        logging.info(f"Successfully exported {len(sheets_data)} sheets to: {file_path}")
        return True
        
    except Exception as e:
        logging.error(f"Error exporting multiple sheets to Excel: {e}")
        return False