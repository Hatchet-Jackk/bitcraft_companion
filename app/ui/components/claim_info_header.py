import logging
import os
import sys
import time
from datetime import datetime

import customtkinter as ctk
import openpyxl
from openpyxl.styles import Alignment, Font
import tkinter as tk
from tkinter import filedialog, messagebox

from app.ui.components.settings_window import SettingsWindow
from app.ui.themes import get_color, register_theme_callback


class ClaimInfoHeader(ctk.CTkFrame):
    """
    Header widget displaying claim information including name, treasury, supplies,
    and time remaining until supplies are depleted based on tile count.
    """

    def __init__(self, master, app, reference_data=None):
        super().__init__(
            master,
            fg_color=get_color("BACKGROUND_SECONDARY"),
            corner_radius=8,
            border_width=1,
            border_color=get_color("BORDER_DEFAULT"),
        )
        self.app = app

        # Register for theme change notifications
        register_theme_callback(self._on_theme_changed)

        # Add claim management attributes
        self.available_claims = []
        self.current_claim_id = None
        self.claim_switching = False

        # Initialize data
        self.claim_name = "Loading..."
        self.treasury = 0
        self.supplies = 0
        self.tile_count = 0
        self.supplies_per_hour = 0
        self.time_remaining = "Calculating..."
        self.traveler_tasks_expiration = None
        self.task_refresh_time = "Unknown"
        self.player_activity_status = "Unknown"
        self._is_initial_loading = True  # Track if we're in initial loading phase

        self.reference_data = reference_data
        self.tile_cost_lookup = {}
        self.has_accurate_tile_costs = False
        self._build_tile_cost_lookup()

        self._create_widgets()

    def _build_tile_cost_lookup(self):
        """Build tile cost lookup dictionary from reference data."""
        try:
            if self.reference_data and "claim_tile_cost" in self.reference_data:
                # Clear existing lookup
                self.tile_cost_lookup = {}

                # Build lookup from reference data
                claim_tile_costs = self.reference_data["claim_tile_cost"]
                for cost_tier in claim_tile_costs:
                    tile_count = cost_tier.get("tile_count")
                    cost_per_tile = cost_tier.get("cost_per_tile")

                    if tile_count is not None and cost_per_tile is not None:
                        self.tile_cost_lookup[tile_count] = cost_per_tile

                self.has_accurate_tile_costs = True
                logging.debug(f"Built tile cost lookup with {len(self.tile_cost_lookup)} tiers")
            else:
                # Fallback to default values if no reference data
                self.tile_cost_lookup = {1: 0.01, 1001: 0.0125}
                self.has_accurate_tile_costs = False
                logging.debug("No claim_tile_cost reference data available, using fallback values")

        except Exception as e:
            logging.error(f"Error building tile cost lookup: {e}")
            # Fallback to default values on error
            self.tile_cost_lookup = {1: 0.01, 1001: 0.0125}
            self.has_accurate_tile_costs = False

    def update_reference_data(self, reference_data):
        """Update reference data and rebuild tile cost lookup."""
        try:
            self.reference_data = reference_data
            self._build_tile_cost_lookup()

            # Recalculate supplies per hour with new data
            if self.tile_count > 0:
                self.supplies_per_hour = self._calculate_supplies_per_hour(self.tile_count)
                self._update_supplies_runout()

        except Exception as e:
            logging.error(f"Error updating reference data in ClaimInfoHeader: {e}")

    def _calculate_supplies_per_hour(self, tile_count):
        """
        Calculate total supplies consumption per hour based on tile count and cost_per_tile.
        """
        if not tile_count or tile_count <= 0:
            return 0.0

        sorted_tile_counts = sorted(self.tile_cost_lookup.keys())
        if tile_count in self.tile_cost_lookup:
            cost_per_tile = self.tile_cost_lookup[tile_count]
            tier_used = tile_count
        elif tile_count <= sorted_tile_counts[0]:
            cost_per_tile = self.tile_cost_lookup[sorted_tile_counts[0]]
            tier_used = sorted_tile_counts[0]
        elif tile_count >= sorted_tile_counts[-1]:
            cost_per_tile = self.tile_cost_lookup[sorted_tile_counts[-1]]
            tier_used = sorted_tile_counts[-1]
        else:
            # Use the largest tile_count less than or equal to the current tile_count
            lower_tiles = max(t for t in sorted_tile_counts if t <= tile_count)
            cost_per_tile = self.tile_cost_lookup[lower_tiles]
            tier_used = lower_tiles

        supplies_per_hour = tile_count * cost_per_tile
        return supplies_per_hour

    def _create_widgets(self):
        """Creates and arranges the header widgets."""
        # Configure grid weights
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=0)

        # Left side - Claim information
        self._create_claim_info_section()

    def _create_claim_info_section(self):
        """Creates the left side claim information display with CTkOptionMenu dropdown."""
        claim_frame = ctk.CTkFrame(self, fg_color="transparent")
        claim_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=10)
        claim_frame.grid_columnconfigure(0, weight=1)

        # Create frame for dropdown and buttons
        dropdown_frame = ctk.CTkFrame(claim_frame, fg_color="transparent")
        dropdown_frame.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 12))

        # Configure grid weights for proper spacing
        dropdown_frame.grid_columnconfigure(0, weight=0)  # Dropdown - fixed size
        dropdown_frame.grid_columnconfigure(1, weight=0)  # Activity - fixed size
        dropdown_frame.grid_columnconfigure(2, weight=0)  # Codex - fixed size
        dropdown_frame.grid_columnconfigure(3, weight=0)  # Settings - fixed size
        dropdown_frame.grid_columnconfigure(4, weight=1)  # Spacer - expandable
        dropdown_frame.grid_columnconfigure(5, weight=0)  # Quit - fixed size, right-aligned

        self.claim_dropdown = ctk.CTkOptionMenu(
            dropdown_frame,
            values=["Loading..."],
            command=self._on_claim_selected,
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=get_color("TEXT_PRIMARY"),
            fg_color=get_color("BACKGROUND_TERTIARY"),
            button_color=get_color("BACKGROUND_PRIMARY"),
            button_hover_color=get_color("BUTTON_HOVER"),
            dropdown_fg_color=get_color("BACKGROUND_TERTIARY"),
            dropdown_hover_color=get_color("BUTTON_HOVER"),
            dropdown_text_color=get_color("TEXT_PRIMARY"),
            corner_radius=8,
            anchor="w",
            state="disabled",  # Start disabled
            height=40,
            width=300,
        )
        self.claim_dropdown.grid(row=0, column=0, sticky="w", padx=(0, 10))

        # Add activity button
        self.activity_button = ctk.CTkButton(
            dropdown_frame,
            text="Activity",
            width=100,
            height=40,
            font=ctk.CTkFont(size=12),
            command=self._open_activity_window,
            fg_color=get_color("TEXT_ACCENT"),
            hover_color=get_color("BUTTON_HOVER"),
            text_color=get_color("TEXT_PRIMARY"),
            corner_radius=8,
            border_width=0,
        )
        self.activity_button.grid(row=0, column=1, sticky="w", padx=(0, 10))

        # Add tooltip to activity button
        self._add_tooltip(self.activity_button, "View recent inventory changes and activity log")

        # Add codex button
        self.codex_button = ctk.CTkButton(
            dropdown_frame,
            text="Codex",
            width=100,
            height=40,
            font=ctk.CTkFont(size=12),
            command=self._open_codex_window,
            fg_color=get_color("TEXT_ACCENT"),
            hover_color=get_color("BUTTON_HOVER"),
            text_color=get_color("TEXT_PRIMARY"),
            corner_radius=8,
            border_width=0,
        )
        self.codex_button.grid(row=0, column=2, sticky="w", padx=(0, 10))

        # Add tooltip to codex button
        self._add_tooltip(self.codex_button, "View material requirements for claim tier advancement")

        # Add settings button
        self.settings_button = ctk.CTkButton(
            dropdown_frame,
            text="Settings",
            width=100,
            height=40,
            font=ctk.CTkFont(size=12),
            command=self._open_settings,
            fg_color=get_color("STATUS_INFO"),
            hover_color=get_color("BUTTON_HOVER"),
            text_color=get_color("TEXT_PRIMARY"),
            corner_radius=8,
            border_width=0,
        )
        self.settings_button.grid(row=0, column=3, sticky="w", padx=(0, 10))

        # Add tooltip to settings button
        self._add_tooltip(self.settings_button, "Open settings window to manage app preferences and data operations")

        # Add quit button
        self.quit_button = ctk.CTkButton(
            dropdown_frame,
            text="Quit",
            width=70,
            height=40,
            font=ctk.CTkFont(size=12),
            command=self._quit_application,
            fg_color=get_color("STATUS_ERROR"),
            hover_color=get_color("STATUS_ERROR"),
            text_color=get_color("TEXT_PRIMARY"),
            corner_radius=8,
        )
        self.quit_button.grid(row=0, column=5, sticky="e")

        # Create info row with treasury, supplies, and supplies run out
        info_frame = ctk.CTkFrame(claim_frame, fg_color="transparent")
        info_frame.grid(row=1, column=0, sticky="w")

        # Treasury with icon
        treasury_frame = self._create_enhanced_info_item(info_frame, "💰 Treasury", "0", get_color("STATUS_WARNING"))
        treasury_frame.grid(row=0, column=0, padx=(0, 20))
        self.treasury_value_label = treasury_frame.winfo_children()[1]

        # Supplies with icon
        supplies_frame = self._create_enhanced_info_item(info_frame, "⚡ Supplies", "0", get_color("STATUS_SUCCESS"))
        supplies_frame.grid(row=0, column=1, padx=(0, 20))
        self.supplies_value_label = supplies_frame.winfo_children()[1]

        # Supplies Run Out with icon
        supplies_runout_frame = self._create_enhanced_info_item(
            info_frame, "⏱️ Depletes In", "Calculating...", get_color("STATUS_WARNING")
        )
        supplies_runout_frame.grid(row=0, column=2, padx=(0, 20))
        self.supplies_runout_label = supplies_runout_frame.winfo_children()[1]

        # Add tooltip to supplies run out label
        self._add_tooltip(self.supplies_runout_label, "This value is approximate and may not exactly match in-game.")

        # Task Refresh with icon
        task_refresh_frame = self._create_enhanced_info_item(info_frame, "🔄 Task Refresh", "Unknown", get_color("STATUS_INFO"))
        task_refresh_frame.grid(row=0, column=3, padx=(0, 20))
        self.task_refresh_label = task_refresh_frame.winfo_children()[1]

        # Add tooltip to task refresh label
        self._add_tooltip(self.task_refresh_label, "Time until traveler tasks refresh with new assignments.")

        # Player Activity Status with icon
        activity_status_frame = self._create_enhanced_info_item(info_frame, "🏃 Activity", "Unknown", get_color("STATUS_INFO"))
        activity_status_frame.grid(row=0, column=4)
        self.activity_status_label = activity_status_frame.winfo_children()[1]

    def _create_enhanced_info_item(self, parent, label_text, value_text, color):
        """Creates an enhanced label-value pair with better styling."""
        frame = ctk.CTkFrame(parent, fg_color="transparent")

        # Label with better styling
        label = ctk.CTkLabel(frame, text=label_text, font=ctk.CTkFont(size=11, weight="normal"), text_color="#b0b0b0")
        label.grid(row=0, column=0, sticky="w")

        # Value with enhanced styling
        value = ctk.CTkLabel(frame, text=value_text, font=ctk.CTkFont(size=14, weight="bold"), text_color=color)
        value.grid(row=1, column=0, sticky="w", pady=(2, 0))

        return frame

    def _add_tooltip(self, widget, text):
        """Adds a tooltip with delay to prevent hover interference."""
        tooltip = None
        tooltip_timer = None

        def show_tooltip():
            nonlocal tooltip
            if tooltip is None:  # Only create if not already created
                try:
                    # Position tooltip away from button edges to prevent interference
                    x = widget.winfo_rootx() + widget.winfo_width() // 2 - 50
                    y = widget.winfo_rooty() + widget.winfo_height() + 5

                    tooltip = tk.Toplevel(widget)
                    tooltip.wm_overrideredirect(True)
                    tooltip.wm_geometry(f"+{x}+{y}")

                    # Ensure tooltip doesn't capture mouse events
                    tooltip.attributes("-topmost", True)
                    tooltip.lift()

                    label = tk.Label(
                        tooltip,
                        text=text,
                        background="#333333",
                        foreground="#ffffff",
                        borderwidth=1,
                        relief="solid",
                        font=("Arial", 9),
                    )
                    label.pack(ipadx=6, ipady=3)
                except Exception:
                    # If tooltip creation fails, just ignore
                    pass

        def on_enter(event):
            nonlocal tooltip_timer
            # Cancel any existing timer
            if tooltip_timer:
                widget.after_cancel(tooltip_timer)
            # Show tooltip after delay to prevent immediate interference
            tooltip_timer = widget.after(400, show_tooltip)

        def on_leave(event):
            nonlocal tooltip, tooltip_timer
            # Cancel pending tooltip
            if tooltip_timer:
                widget.after_cancel(tooltip_timer)
                tooltip_timer = None
            # Destroy existing tooltip
            if tooltip:
                try:
                    tooltip.destroy()
                except Exception:
                    pass
                tooltip = None

        widget.bind("<Enter>", on_enter)
        widget.bind("<Leave>", on_leave)

    def _on_claim_selected(self, selected_claim_name: str):
        """
        Handles claim selection from the CTkOptionMenu dropdown.
        """
        if self.claim_switching:
            return

        # Find the claim ID from the selected name (handle both key formats)
        for claim in self.available_claims:
            claim_name = claim.get("claim_name") or claim.get("name")
            if claim_name == selected_claim_name:
                claim_id = claim.get("claim_id") or claim.get("entity_id")

                # Only switch if it's different from current
                if claim_id != self.current_claim_id:
                    logging.info(f"User selected claim: {selected_claim_name} ({claim_id})")

                    # Start claim switching
                    self.set_claim_switching(True, f"Switching to {selected_claim_name}...")

                    # Notify the main app to perform the switch
                    if hasattr(self.app, "switch_to_claim"):
                        self.app.switch_to_claim(claim_id)
                    else:
                        logging.error("Main app does not support claim switching")
                        self.set_claim_switching(False)
                break

    def update_claim_data(self, claim_data):
        """
        Updates the header with new claim information.
        """
        try:
            # Update stored values
            self.claim_name = claim_data.get("name", "Unknown Claim")
            self.treasury = claim_data.get("treasury", 0)
            self.supplies = claim_data.get("supplies", 0)
            self.tile_count = claim_data.get("tile_count", 0)

            # Calculate supplies per hour based on tile count
            self.supplies_per_hour = self._calculate_supplies_per_hour(self.tile_count)

            # Update UI labels
            if not self.claim_switching:
                self._update_dropdown_selection()

            self.treasury_value_label.configure(text=f"{self.treasury:,}")
            self.supplies_value_label.configure(text=f"{self.supplies:,}")

            # Calculate and update supplies run out time
            self._update_supplies_runout()

            # Notify MainWindow that claim info data loading completed (for loading overlay detection)
            if hasattr(self.app, "is_loading") and self.app.is_loading:
                if hasattr(self.app, "received_data_types"):
                    self.app.received_data_types.add("claim_info")
                    logging.info(f"[ClaimInfoHeader] Notified MainWindow of claim info data completion")
                    if hasattr(self.app, "_check_all_data_loaded"):
                        self.app._check_all_data_loaded()

        except Exception as e:
            logging.error(f"Error updating claim data in header: {e}")

    def update_player_activity_status(self, activity_data):
        """
        Updates the player activity status display.

        Args:
            activity_data: Dict containing activity status information
        """
        try:
            status = activity_data.get("status", "Unknown")
            stamina_percentage = activity_data.get("stamina_percentage", 0.0)

            # Update stored status
            self.player_activity_status = status

            # Choose color based on activity status
            status_colors = {
                "Active": "#FF9800",  # Orange - player is performing actions
                "Resting": "#FFC107",  # Amber - stamina recovering
                "Idle": "#4CAF50",  # Green - at full stamina
                "Unknown": "#9E9E9E",  # Gray - status unknown
            }

            color = status_colors.get(status, "#9E9E9E")

            # Update the display label
            display_text = f"{status}"
            if status in ["Active", "Resting"] and stamina_percentage > 0:
                display_text += f" ({stamina_percentage:.0f}%)"

            self.activity_status_label.configure(text=display_text, text_color=color)

            logging.debug(f"[ClaimInfoHeader] Updated activity status: {status} ({stamina_percentage:.1f}%)")

        except Exception as e:
            logging.error(f"Error updating player activity status: {e}")
            self.activity_status_label.configure(text="Error", text_color="#f44336")

    def _update_supplies_runout(self):
        """Calculates and updates the time until supplies run out based on tile count."""
        try:
            # Show loading state if we don't have accurate tile cost data yet
            if not self.has_accurate_tile_costs:
                self.time_remaining = "Loading..."
                color = "#cccccc"
                self.supplies_runout_label.configure(text=self.time_remaining, text_color=color)
                return

            if self.supplies_per_hour <= 0:
                self.time_remaining = "N/A"
                color = "#cccccc"
            elif self.supplies <= 0:
                self.time_remaining = "Depleted"
                color = "#f44336"
            else:
                total_seconds = int(self.supplies * 3600 / self.supplies_per_hour)
                days = total_seconds // 86400
                hours = (total_seconds % 86400) // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60

                if total_seconds < 1800:  # less than 30 minutes
                    self.time_remaining = f"{minutes}m {seconds}s"
                    color = "#f44336"  # Red for urgent
                elif days == 0 and hours == 0:
                    self.time_remaining = f"{minutes}m"
                    color = "#f44336"
                elif days == 0:
                    self.time_remaining = f"{hours}h {minutes}m"
                    color = "#FF9800"  # Orange for warning
                elif days < 7:
                    self.time_remaining = f"{days}d {hours}h {minutes}m"
                    color = "#FFC107"  # Amber for caution
                else:
                    self.time_remaining = f"{days}d {hours}h {minutes}m"
                    color = "#4CAF50"  # Green for safe

            # Update the label with appropriate color
            self.supplies_runout_label.configure(text=self.time_remaining, text_color=color)

        except Exception as e:
            logging.error(f"Error calculating supplies run out time: {e}")
            self.time_remaining = "Error"
            self.supplies_runout_label.configure(text=self.time_remaining, text_color="#f44336")

    def refresh_supplies_runout(self):
        """Manually refresh the supplies run out calculation (for periodic updates)."""
        self._update_supplies_runout()

    def _open_settings(self):
        """Opens the settings window."""
        try:
            # Create and show the settings window
            settings_window = SettingsWindow(self.app, self.app)

            logging.info("Settings window opened from header")

        except Exception as e:
            logging.error(f"Error opening settings window: {e}")
            messagebox.showerror("Settings Error", f"Failed to open settings:\n{str(e)}")

    def _open_activity_window(self):
        """Opens the activity logs window."""
        try:
            if hasattr(self.app, "_open_activity_window"):
                self.app._open_activity_window()
            else:
                logging.warning("Main app does not have _open_activity_window method")

        except Exception as e:
            logging.error(f"Error opening activity window from header: {e}")

    def _open_codex_window(self):
        """Opens the codex material requirements window."""
        try:
            if hasattr(self.app, "_open_codex_window"):
                self.app._open_codex_window()
            else:
                logging.warning("Main app does not have _open_codex_window method")

        except Exception as e:
            logging.error(f"Error opening codex window from header: {e}")

    def _reset_refresh_button(self):
        """Reset the refresh button to its normal state."""
        # This method is no longer needed since refresh button was removed
        pass

    def update_available_claims(self, claims_list: list, current_claim_id: str = None):
        """
        Updates the list of available claims using CTkOptionMenu.
        """
        self.available_claims = claims_list
        if current_claim_id:
            self.current_claim_id = current_claim_id

        # Extract claim names for the dropdown (handle both key formats)
        claim_names = [claim.get("claim_name") or claim.get("name", "Unknown Claim") for claim in claims_list]

        # Update dropdown state and values based on available claims
        if len(claims_list) > 1:
            # Multiple claims - enable dropdown
            self.claim_dropdown.configure(values=claim_names, state="normal")

            # Set current selection (handle both key formats)
            current_claim_name = None
            for claim in claims_list:
                claim_entity_id = claim.get("entity_id") or claim.get("claim_id")
                if claim_entity_id == current_claim_id:
                    current_claim_name = claim.get("claim_name") or claim.get("name")
                    break

            if current_claim_name:
                self.claim_dropdown.set(current_claim_name)
                self.claim_name = current_claim_name

        elif len(claims_list) == 1:
            # Single claim - show name but disable interaction
            single_claim = claims_list[0]
            claim_name = single_claim.get("claim_name") or single_claim.get("name", "Unknown Claim")
            self.claim_dropdown.configure(values=[claim_name], state="disabled")
            self.claim_dropdown.set(claim_name)
            self.current_claim_id = single_claim.get("entity_id") or single_claim.get("claim_id")
            self.claim_name = claim_name

        else:
            # No claims - show error state
            self.claim_dropdown.configure(values=["No Claims Available"], state="disabled")
            self.claim_dropdown.set("No Claims Available")

        logging.info(f"Updated available claims: {len(claims_list)} claims, current: {current_claim_id}")

    def _update_dropdown_selection(self):
        """Updates the dropdown to show the current claim name."""
        if self.available_claims and not self.claim_switching:
            # Find current claim name and update dropdown (handle both key formats)
            for claim in self.available_claims:
                claim_entity_id = claim.get("entity_id") or claim.get("claim_id")
                if claim_entity_id == self.current_claim_id:
                    claim_name = claim.get("claim_name") or claim.get("name")
                    self.claim_dropdown.set(claim_name)
                    break

    def set_claim_switching(self, switching: bool, message: str = ""):
        """
        Sets the claim switching state. Instead of showing loading in dropdown,
        this signals the main app to show the loading overlay.
        """
        self.claim_switching = switching

        if switching:
            # Disable dropdown during switching
            self.claim_dropdown.configure(state="disabled")

            # Notify main app to show loading overlay with custom message
            if hasattr(self.app, "show_loading_with_message"):
                self.app.show_loading_with_message(message)

        else:
            # Re-enable dropdown and restore normal state
            if len(self.available_claims) > 1:
                self.claim_dropdown.configure(state="normal")
            else:
                self.claim_dropdown.configure(state="disabled")

            # Notify main app to hide loading overlay
            if hasattr(self.app, "hide_loading"):
                self.app.hide_loading()

            # Restore proper selection
            self._update_dropdown_selection()

    def handle_claim_switch_complete(self, claim_id: str, claim_name: str):
        """
        Handles the completion of a claim switch operation.
        """
        self.current_claim_id = claim_id
        self.claim_name = claim_name

        # End switching state
        self.set_claim_switching(False)

        # Update dropdown selection
        self._update_dropdown_selection()

        logging.info(f"Claim switch completed: {claim_name}")

    def handle_claim_switch_error(self, error_message: str):
        """
        Handles errors during claim switching.
        """
        # End switching state
        self.set_claim_switching(False)

        logging.error(f"Claim switch error: {error_message}")

    def get_claims_summary(self) -> str:
        """Returns a summary of available claims for debugging."""
        if not self.available_claims:
            return "No claims available"

        current_name = "None"
        if self.current_claim_id:
            for claim in self.available_claims:
                if claim["claim_id"] == self.current_claim_id:
                    current_name = claim["claim_name"]
                    break

        return f"{len(self.available_claims)} claims, current: {current_name}, switching: {self.claim_switching}"

    def initialize_with_claims(self, claims_list: list, current_claim_id: str = None):
        """
        Initializes the header with the claims list on first load.
        """
        try:
            self.update_available_claims(claims_list, current_claim_id)

            # If we have a current claim, update the display
            if current_claim_id:
                current_claim = None
                for claim in claims_list:
                    if claim["entity_id"] == current_claim_id:
                        current_claim = claim
                        break

                if current_claim:
                    # Update header with initial claim data
                    self.update_claim_data(
                        {
                            "name": current_claim["name"],
                            "treasury": current_claim.get("treasury", 0),
                            "supplies": current_claim.get("supplies", 0),
                            "tile_count": current_claim.get("tile_count", 0),
                        }
                    )

            logging.info(f"Header initialized with {len(claims_list)} claims")

        except Exception as e:
            logging.error(f"Error initializing header with claims: {e}")

    def update_task_refresh_expiration(
        self,
        expiration_time: int,
        is_initial_subscription: bool = False,
        source: str = "unknown",
    ):
        """
        Updates the traveler tasks expiration time and starts countdown.

        Args:
            expiration_time: Seconds since Unix epoch when tasks refresh
            is_initial_subscription: True if this data came from InitialSubscription
            source: Source of the update (subscription, transaction)
        """
        try:
            logging.debug(
                f"[ClaimInfoHeader] Timer update: expiration={expiration_time}, is_initial={is_initial_subscription}, source={source}"
            )

            # Update expiration time - always accept the new value
            self.traveler_tasks_expiration = expiration_time

            # Mark the source of this update for timer logic
            self._last_update_source = source
            self._from_initial_subscription = is_initial_subscription

            # Mark that we've received initial data
            if is_initial_subscription:
                self._is_initial_loading = False

            self._update_task_refresh_countdown()

            # Start periodic updates if not already running
            if not hasattr(self, "_task_refresh_timer_running"):
                self._task_refresh_timer_running = True
                self._schedule_task_refresh_update()

        except Exception as e:
            logging.error(f"Error updating task refresh expiration: {e}")

    def _update_task_refresh_countdown(self):
        """Calculates and updates the task refresh countdown."""
        try:
            if self.traveler_tasks_expiration is None:
                self.task_refresh_time = "Loading..."
                color = "#9C27B0"
            else:
                current_time_seconds = time.time()
                time_diff_seconds = self.traveler_tasks_expiration - current_time_seconds

                if time_diff_seconds <= 0:
                    # Timer expired
                    current_dt = datetime.fromtimestamp(current_time_seconds)
                    expiration_dt = datetime.fromtimestamp(self.traveler_tasks_expiration)

                    # Check how long we've been in expired state
                    expired_duration = abs(time_diff_seconds)

                    if expired_duration < 60:  # Less than 1 minute - show refreshing
                        self.task_refresh_time = "Refreshing..."
                        color = "#FF9800"  # Orange for activity/refreshing state
                    elif self._is_initial_loading:
                        # During initial loading, show waiting for server
                        if expired_duration < 300:  # Less than 5 minutes during initial load
                            logging.debug(f"Initial loading - timer expired {expired_duration:.1f}s ago, waiting for server")
                            self.task_refresh_time = "Waiting for server..."
                            color = "#9E9E9E"  # Gray for waiting/inactive state
                        else:
                            # Initial loading taking too long, show error
                            logging.warning(f"Initial timer load taking over 5 minutes")
                            self.task_refresh_time = "Connection issue"
                            color = "#f44336"  # Red for error
                    else:
                        # App is running normally - auto-reset to 4 hours
                        logging.info(
                            f"Timer expired during normal operation, auto-resetting to 4 hours (expired_for={expired_duration:.1f}s)"
                        )
                        four_hours_from_now = current_time_seconds + (4 * 60 * 60)
                        self.traveler_tasks_expiration = four_hours_from_now

                        # Recalculate with new expiration
                        time_diff_seconds = self.traveler_tasks_expiration - current_time_seconds
                        # Fall through to normal countdown logic below

                    # Only return early if we're not auto-resetting
                    if time_diff_seconds <= 0:
                        self.task_refresh_label.configure(text=self.task_refresh_time, text_color=color)
                        return

                # Continue with normal countdown display logic...
                total_seconds = int(time_diff_seconds)
                days = total_seconds // 86400
                hours = (total_seconds % 86400) // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60

                # Format display based on time remaining - show largest relevant units
                if total_seconds < 60:  # less than 1 minute - show seconds only
                    self.task_refresh_time = f"{seconds}s"
                    color = "#FF5722"  # Red for very soon
                elif total_seconds < 1800:  # less than 30 minutes - show minutes and seconds
                    self.task_refresh_time = f"{minutes}m {seconds}s"
                    color = "#FF9800"  # Orange for soon
                elif total_seconds < 3600:  # less than 1 hour - show minutes only
                    self.task_refresh_time = f"{minutes}m"
                    color = "#FF9800"  # Orange for soon
                elif days == 0 and hours > 0:  # same day with hours - show hours and minutes
                    self.task_refresh_time = f"{hours}h {minutes}m"
                    color = "#FFC107"  # Amber for today
                elif days > 0:  # more than a day - show days and hours
                    self.task_refresh_time = f"{days}d {hours}h"
                    color = "#9C27B0"  # Purple for future
                else:  # fallback
                    self.task_refresh_time = f"{hours}h {minutes}m"
                    color = "#FFC107"

            # Update the label with appropriate color
            self.task_refresh_label.configure(text=self.task_refresh_time, text_color=color)

        except Exception as e:
            logging.error(f"Error calculating task refresh countdown: {e}")
            self.task_refresh_time = "Error"
            self.task_refresh_label.configure(text=self.task_refresh_time, text_color="#f44336")

    def _schedule_task_refresh_update(self):
        """Schedules the next task refresh countdown update."""
        try:
            # Update every second for accurate countdown
            self.after(1000, self._periodic_task_refresh_update)
        except Exception as e:
            logging.error(f"Error scheduling task refresh update: {e}")

    def _periodic_task_refresh_update(self):
        """Periodic update method for task refresh countdown."""
        try:
            if hasattr(self, "_task_refresh_timer_running") and self._task_refresh_timer_running:
                self._update_task_refresh_countdown()
                self._schedule_task_refresh_update()
        except Exception as e:
            logging.error(f"Error in periodic task refresh update: {e}")

    def stop_task_refresh_timer(self):
        """Stops the task refresh countdown timer."""
        try:
            if hasattr(self, "_task_refresh_timer_running"):
                self._task_refresh_timer_running = False
        except Exception as e:
            logging.error(f"Error stopping task refresh timer: {e}")

    def update_task_refresh_retry_status(self, status, message, delay=0):
        """
        Update task refresh display with retry status information.

        Args:
            status: "retrying" or "failed"
            message: Display message for the user
            delay: Retry delay in seconds (for countdown)
        """
        try:
            if status == "retrying":
                self.task_refresh_time = message
                color = get_color("STATUS_WARNING")

                # Start countdown for retry delay if delay > 0
                if delay > 0:
                    self._start_retry_countdown(delay)

            elif status == "failed":
                self.task_refresh_time = "Connection failed"
                color = get_color("STATUS_ERROR")

            # Update the label
            self.task_refresh_label.configure(text=self.task_refresh_time, text_color=color)
            logging.debug(f"Task refresh retry status updated: {status} - {message}")

        except Exception as e:
            logging.error(f"Error updating task refresh retry status: {e}")

    def _start_retry_countdown(self, delay):
        """Start a countdown timer for retry delay."""
        try:
            self._retry_countdown_remaining = delay
            self._update_retry_countdown()

        except Exception as e:
            logging.error(f"Error starting retry countdown: {e}")

    def _update_retry_countdown(self):
        """Update retry countdown display."""
        try:
            if hasattr(self, "_retry_countdown_remaining") and self._retry_countdown_remaining > 0:
                self.task_refresh_time = f"Retrying in {self._retry_countdown_remaining}s..."
                self.task_refresh_label.configure(text=self.task_refresh_time, text_color=get_color("STATUS_WARNING"))

                self._retry_countdown_remaining -= 1

                # Schedule next update
                self.after(1000, self._update_retry_countdown)
            elif hasattr(self, "_retry_countdown_remaining"):
                # Countdown finished
                self.task_refresh_time = "Retrying now..."
                self.task_refresh_label.configure(text=self.task_refresh_time, text_color=get_color("STATUS_WARNING"))

        except Exception as e:
            logging.error(f"Error updating retry countdown: {e}")

    def _request_fresh_player_state(self):
        """
        Request fresh player_state data to get updated traveler_tasks_expiration.
        This is called when the timer hits 0 to detect task refresh.
        """
        try:
            # Access the data service through the main app to request fresh data
            if hasattr(self.app, "data_service") and self.app.data_service:

                # As a fallback, schedule a check in 30 seconds to see if data has updated
                self.after(30000, self._check_for_updated_expiration)
            else:
                logging.warning("Cannot request fresh player state - no data service available")
        except Exception as e:
            logging.error(f"Error requesting fresh player state: {e}")

    def _check_for_updated_expiration(self):
        """
        Check if the expiration time has been updated (indicating tasks refreshed).
        This is a fallback method when we can't directly request fresh data.
        """
        try:
            # This method is called periodically to check if we need to refresh player state
            # when tasks might be refreshing (no longer needed with improved logic)
            pass
        except Exception as e:
            logging.error(f"Error checking for updated expiration: {e}")

    def _export_data(self):
        """Exports all table data to an Excel file using native file dialog."""
        try:
            # Generate default filename
            # Try multiple sources for the claim name
            dropdown_value = self.claim_dropdown.get() if hasattr(self, "claim_dropdown") else None

            # Try to find claim name from available claims using current claim ID
            claims_name = None
            if self.current_claim_id and hasattr(self, "available_claims"):
                for claim in self.available_claims:
                    claim_entity_id = claim.get("entity_id") or claim.get("claim_id")
                    if str(claim_entity_id) == str(self.current_claim_id):
                        claims_name = claim.get("claim_name") or claim.get("name")
                        break

            # Choose the best available claim name
            if self.claim_name and self.claim_name not in ["Loading...", "Unknown Claim"]:
                claim_name = self.claim_name
                source = "self.claim_name"
            elif dropdown_value and dropdown_value not in ["Loading...", "Unknown Claim"]:
                claim_name = dropdown_value
                source = "dropdown"
            elif claims_name and claims_name not in ["Loading...", "Unknown Claim"]:
                claim_name = claims_name
                source = "available_claims"
                # Update self.claim_name for consistency
                self.claim_name = claims_name
            else:
                claim_name = "Unknown_Claim"
                source = "fallback"

            logging.info(
                f"Export claim name source: {source} -> '{claim_name}' (self.claim_name: '{self.claim_name}', dropdown: '{dropdown_value}', claims_name: '{claims_name}')"
            )

            claim_name = claim_name.replace(" ", "_").replace("/", "-")
            claim_id = self.current_claim_id if self.current_claim_id else "unknown"
            date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"{claim_name}_{claim_id}_{date_str}.xlsx"

            # Show file save dialog
            file_path = filedialog.asksaveasfilename(
                title="Export Claim Data",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=default_filename,
            )

            if not file_path:
                logging.info("Export cancelled by user")
                return

            # Create workbook
            workbook = openpyxl.Workbook()

            # Remove default sheet
            default_sheet = workbook.active
            workbook.remove(default_sheet)

            # Export each tab's data
            tabs_data = self._get_all_tabs_data()

            for tab_name, tab_data in tabs_data.items():
                if tab_data:
                    self._create_excel_sheet(workbook, tab_name, tab_data)

            # Add claim info sheet
            self._create_claim_info_sheet(workbook)

            # Save the workbook
            workbook.save(file_path)

            # Show success message
            messagebox.showinfo("Export Successful", f"Data exported successfully to:\n{os.path.basename(file_path)}")

            logging.info(f"Data exported to: {file_path}")

        except Exception as e:
            logging.error(f"Error exporting data: {e}")
            messagebox.showerror("Export Error", f"Failed to export data:\n{str(e)}")

    def _get_all_tabs_data(self):
        """Retrieves data from all tabs in the main application."""
        tabs_data = {}

        try:
            if hasattr(self.app, "tabs"):
                for tab_name, tab_instance in self.app.tabs.items():
                    if hasattr(tab_instance, "filtered_data"):
                        # Get filtered data from each tab
                        data = tab_instance.filtered_data
                        if data:
                            tabs_data[tab_name] = {"headers": getattr(tab_instance, "headers", []), "data": data}

        except Exception as e:
            logging.error(f"Error getting tabs data: {e}")

        return tabs_data

    def _create_excel_sheet(self, workbook, tab_name, tab_data):
        """Creates an Excel sheet for a specific tab's data."""
        try:
            sheet = workbook.create_sheet(title=tab_name)
            headers = tab_data.get("headers", [])
            data = tab_data.get("data", [])

            if not headers or not data:
                return

            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    # Handle special cases for different tab structures
                    if tab_name == "Traveler's Tasks":
                        value = self._get_traveler_task_value(row_data, header)
                    else:
                        # Map header names to actual data field names
                        value = self._get_mapped_value(row_data, header, tab_name)

                    # Convert to string for Excel
                    sheet.cell(row=row_idx, column=col_idx, value=str(value))

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        except Exception as e:
            logging.error(f"Error creating Excel sheet for {tab_name}: {e}")

    def _get_mapped_value(self, row_data, header, tab_name):
        """Maps header names to actual data field names for different tabs."""
        try:
            # Handle dataclass objects by converting to dict if needed
            if hasattr(row_data, "to_dict"):
                row_dict = row_data.to_dict()
            elif hasattr(row_data, "__dict__"):
                row_dict = row_data.__dict__
            else:
                row_dict = row_data

            # Define header to field mappings for different tabs
            header_mappings = {
                "Claim Inventory": {
                    "Item": "name",
                    "Tier": "tier",
                    "Quantity": "quantity",
                    "Tag": "tag",
                    "Containers": "containers",
                },
                "Passive Crafting": {
                    "Item": "item_name",
                    "Quantity": "quantity",
                    "Tier": "tier",
                    "Tag": "tag",
                    "Building": "building_name",
                    "Recipe": "recipe_name",
                    "Progress": "progress_percentage",
                    "Time Left": "time_remaining",
                },
                "Active Crafting": {
                    "Item": "item_name",
                    "Quantity": "quantity",
                    "Tier": "tier",
                    "Tag": "tag",
                    "Building": "building_name",
                    "Recipe": "recipe_name",
                    "Status": "status",
                },
            }

            # Get the mapping for this tab, default to generic mapping
            tab_mapping = header_mappings.get(tab_name, {})
            field_name = tab_mapping.get(header)

            if field_name:
                # Use the mapped field name
                value = row_dict.get(field_name, "")
            else:
                # Fallback to generic header mapping
                header_key = header.lower().replace(" ", "_").replace("'", "")
                value = row_dict.get(header_key, row_dict.get(header.lower(), ""))

            # Special handling for containers field
            if header == "Containers" and isinstance(value, dict):
                # Convert containers dict to readable format
                if value:
                    container_list = [f"{name}: {qty}" for name, qty in value.items()]
                    return "; ".join(container_list)
                else:
                    return "None"

            return value

        except Exception as e:
            logging.error(f"Error getting mapped value for header '{header}': {e}")
            return ""

    def _get_traveler_task_value(self, row_data, header):
        """Gets the correct value for traveler tasks data structure."""
        try:
            if header == "Traveler":
                return row_data.get("traveler", "")
            elif header == "Item":
                # For parent rows, show completion summary
                operations = row_data.get("operations", [])
                if operations:
                    completed = row_data.get("completed_count", 0)
                    total = row_data.get("total_count", 0)
                    return f"Tasks ({completed}/{total} completed)"
                return row_data.get("item", "")
            elif header == "Completed":
                return row_data.get("completed", "")
            elif header == "Status":
                return row_data.get("status", "")
            else:
                # For other fields, check operations if main row doesn't have it
                value = row_data.get(header.lower().replace(" ", "_"), "")
                if not value:
                    operations = row_data.get("operations", [])
                    if operations:
                        # Get from first operation
                        return operations[0].get(header.lower().replace(" ", "_"), "")
                return value
        except Exception as e:
            logging.error(f"Error getting traveler task value: {e}")
            return ""

    def _create_claim_info_sheet(self, workbook):
        """Creates a sheet with claim information."""
        try:
            sheet = workbook.create_sheet(title="Claim Info")

            # Claim information
            claim_info = [
                ["Claim Name", self.claim_name],
                ["Claim ID", self.current_claim_id or "Unknown"],
                ["Treasury", f"{self.treasury:,}"],
                ["Supplies", f"{self.supplies:,}"],
                ["Tile Count", str(self.tile_count)],
                ["Supplies per Hour", f"{self.supplies_per_hour:.3f}"],
                ["Time Until Depletion", self.time_remaining],
                ["Task Refresh Time", self.task_refresh_time],
                ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ]

            for row_idx, (label, value) in enumerate(claim_info, 1):
                sheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                sheet.cell(row=row_idx, column=2, value=str(value))

            # Auto-adjust column widths
            sheet.column_dimensions["A"].width = 20
            sheet.column_dimensions["B"].width = 30

        except Exception as e:
            logging.error(f"Error creating claim info sheet: {e}")

    def _quit_application(self):
        """Quits the application."""
        try:
            logging.info("User initiated application quit from header")

            # Call the main window's closing method
            if hasattr(self.app, "on_closing"):
                self.app.on_closing()
            else:
                # Fallback: try to close the root window
                self.app.quit()

        except Exception as e:
            logging.error(f"Error quitting application: {e}")
            try:
                self.app.quit()
            except:
                sys.exit(0)

    def _on_theme_changed(self, old_theme: str, new_theme: str):
        """Handle theme change by updating colors."""
        try:
            # Update main frame background
            self.configure(fg_color=get_color("BACKGROUND_SECONDARY"))

            # Update claim dropdown if it exists
            if hasattr(self, "claim_dropdown"):
                self.claim_dropdown.configure(
                    text_color=get_color("TEXT_PRIMARY"),
                    fg_color=get_color("BACKGROUND_TERTIARY"),
                    button_color=get_color("BACKGROUND_PRIMARY"),
                    button_hover_color=get_color("BUTTON_HOVER"),
                    dropdown_fg_color=get_color("BACKGROUND_TERTIARY"),
                    dropdown_hover_color=get_color("BUTTON_HOVER"),
                    dropdown_text_color=get_color("TEXT_PRIMARY"),
                )

            # Update activity button if it exists
            if hasattr(self, "activity_button"):
                self.activity_button.configure(
                    fg_color=get_color("TEXT_ACCENT"), hover_color=get_color("BUTTON_HOVER"), text_color=get_color("TEXT_PRIMARY")
                )

            # Update codex button if it exists
            if hasattr(self, "codex_button"):
                self.codex_button.configure(
                    fg_color=get_color("TEXT_ACCENT"), hover_color=get_color("BUTTON_HOVER"), text_color=get_color("TEXT_PRIMARY")
                )

            # Update settings button if it exists
            if hasattr(self, "settings_button"):
                self.settings_button.configure(
                    fg_color=get_color("STATUS_INFO"), hover_color=get_color("BUTTON_HOVER"), text_color=get_color("TEXT_PRIMARY")
                )

            # Update quit button if it exists
            if hasattr(self, "quit_button"):
                self.quit_button.configure(
                    fg_color=get_color("STATUS_ERROR"),
                    hover_color=get_color("STATUS_ERROR"),
                    text_color=get_color("TEXT_PRIMARY"),
                )

            logging.debug(f"Claim info header theme changed from {old_theme} to {new_theme}")

        except Exception as e:
            logging.error(f"Error updating claim info header theme: {e}")
